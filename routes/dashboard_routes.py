import re
from flask import Blueprint, render_template, jsonify, request, send_file, current_app
from flask_login import login_required, current_user
from services import kpi_service
from datetime import datetime
from zoneinfo import ZoneInfo

from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

dashboard_bp = Blueprint("dashboard", __name__)


@dashboard_bp.before_request
def proteger_apis():
    if request.path.startswith("/api/") and not current_user.is_authenticated:
        return jsonify({"error": "Sesión expirada o usuario no autenticado."}), 401


# =========================
# PÁGINAS DEL SISTEMA
# =========================

@dashboard_bp.route("/")
@dashboard_bp.route("/dashboard")
@login_required
def dashboard():
    return render_template("dashboard.html")


@dashboard_bp.route("/atenciones")
@login_required
def atenciones():
    return render_template("atenciones.html")


@dashboard_bp.route("/valorizacion")
@login_required
def valorizacion():
    return render_template("valorizacion.html")


@dashboard_bp.route("/calidad")
@login_required
def calidad():
    return render_template("calidad.html")


@dashboard_bp.route("/etl")
@login_required
def etl():
    if current_user.rol != "admin":
        return render_template("sin_permiso.html"), 403

    return render_template("etl.html")


# =========================
# APIS
# =========================

def _error_response(message="No se pudo completar la solicitud.", status_code=500):
    current_app.logger.exception(message)
    return jsonify({"error": message}), status_code


@dashboard_bp.route("/api/resumen")
def api_resumen():
    try:
        data = kpi_service.obtener_resumen_general()
        return jsonify(data)
    except Exception:
        return _error_response()


@dashboard_bp.route("/api/atenciones-mes")
def api_atenciones_mes():
    try:
        data = kpi_service.obtener_atenciones_mes()
        return jsonify(data)
    except Exception:
        return _error_response()


@dashboard_bp.route("/api/valorizacion-mes")
def api_valorizacion_mes():
    try:
        data = kpi_service.obtener_valorizacion_mes()
        return jsonify(data)
    except Exception:
        return _error_response()


@dashboard_bp.route("/api/establecimientos")
def api_establecimientos():
    try:
        data = kpi_service.obtener_establecimientos()
        return jsonify(data)
    except Exception:
        return _error_response()


@dashboard_bp.route("/api/servicios")
def api_servicios():
    try:
        data = kpi_service.obtener_servicios()
        return jsonify(data)
    except Exception:
        return _error_response()


@dashboard_bp.route("/api/sexo")
def api_sexo():
    try:
        data = kpi_service.obtener_sexo()
        return jsonify(data)
    except Exception:
        return _error_response()


@dashboard_bp.route("/api/observados")
def api_observados():
    try:
        data = kpi_service.obtener_observados()
        return jsonify(data)
    except Exception:
        return _error_response()


@dashboard_bp.route("/api/estado-valorizacion")
def api_estado_valorizacion():
    try:
        data = kpi_service.obtener_estado_valorizacion()
        return jsonify(data)
    except Exception:
        return _error_response()


@dashboard_bp.route("/api/calidad")
def api_calidad():
    try:
        data = kpi_service.obtener_calidad_datos()
        return jsonify(data)
    except Exception:
        return _error_response()


@dashboard_bp.route("/api/componentes-valorizacion")
def api_componentes_valorizacion():
    try:
        data = kpi_service.obtener_componentes_valorizacion()
        return jsonify(data)
    except Exception:
        return _error_response()

@dashboard_bp.route("/api/lista-servicios")
@login_required
def api_lista_servicios():
    try:
        data = kpi_service.obtener_lista_servicios()
        return jsonify(data)
    except Exception:
        return _error_response()


@dashboard_bp.route("/api/lista-establecimientos")
@login_required
def api_lista_establecimientos():
    try:
        return jsonify(kpi_service.obtener_lista_establecimientos())
    except Exception:
        return _error_response()


@dashboard_bp.route("/api/atenciones-servicio-tiempo")
@login_required
def api_atenciones_servicio_tiempo():
    try:
        servicio = request.args.get("servicio", "TODOS")
        establecimiento = request.args.get("establecimiento", "TODOS")
        granularidad = request.args.get("granularidad", "mes")
        fecha_inicio = request.args.get("fecha_inicio")
        fecha_fin = request.args.get("fecha_fin")

        if fecha_inicio == "":
            fecha_inicio = None

        if fecha_fin == "":
            fecha_fin = None

        data = kpi_service.obtener_atenciones_servicio_tiempo(
            servicio=servicio,
            establecimiento=establecimiento,
            granularidad=granularidad,
            fecha_inicio=fecha_inicio,
            fecha_fin=fecha_fin
        )

        return jsonify(data)

    except Exception as e:
        return jsonify({"error": str(e)}), 500
    
@dashboard_bp.route("/api/exportar-atenciones-excel")
@login_required
def exportar_atenciones_excel():

    try:
        servicio = request.args.get("servicio", "TODOS")
        establecimiento = request.args.get("establecimiento", "TODOS")
        granularidad = request.args.get("granularidad", "mes")
        fecha_inicio = request.args.get("fecha_inicio")
        fecha_fin = request.args.get("fecha_fin")


        data = kpi_service.obtener_atenciones_servicio_tiempo(
            servicio=servicio,
            establecimiento=establecimiento,
            granularidad=granularidad,
            fecha_inicio=fecha_inicio,
            fecha_fin=fecha_fin
        )


        # ==============================
        # OBTENER INFORMACIÓN DEL SERVICIO
        # ==============================

        codigo_servicio = "TODOS"
        nombre_servicio = "Todos los servicios"


        if servicio != "TODOS":

            lista_servicios = kpi_service.obtener_lista_servicios()

            servicio_info = next(
                (
                    item for item in lista_servicios
                    if str(item["cod_servicio"]).strip() == str(servicio).strip()
                ),
                None
            )


            if servicio_info:
                codigo_servicio = servicio_info["cod_servicio"]
                nombre_servicio = servicio_info["descripcion_servicio"]

        codigo_establecimiento = "TODOS"
        nombre_establecimiento = "Todos los establecimientos"
        if establecimiento != "TODOS":
            establecimiento_info = next(
                (item for item in kpi_service.obtener_lista_establecimientos()
                 if str(item["codigo_eess"]).strip() == str(establecimiento).strip()),
                None,
            )
            if establecimiento_info:
                codigo_establecimiento = establecimiento_info["codigo_eess"]
                nombre_establecimiento = establecimiento_info["nombre_eess"]



        # ==============================
        # CREAR EXCEL
        # ==============================

        wb = Workbook()

        ws = wb.active
        ws.title = "Atenciones"


        # Título

        ws["A1"] = "Reporte de atenciones SIS"

        ws["A1"].font = Font(
            bold=True,
            size=14,
            color="FFFFFF"
        )

        ws["A1"].fill = PatternFill(
            "solid",
            fgColor="17365D"
        )

        ws.merge_cells("A1:D1")



        # ==============================
        # INFORMACIÓN DEL FILTRO
        # ==============================

        ws["A3"] = "Código servicio"
        ws["B3"] = codigo_servicio


        ws["A4"] = "Servicio"
        ws["B4"] = nombre_servicio
        ws["C3"] = "Código establecimiento"
        ws["D3"] = codigo_establecimiento
        ws["C4"] = "Establecimiento"
        ws["D4"] = nombre_establecimiento
        ws["C3"].font = Font(bold=True)
        ws["C4"].font = Font(bold=True)


        ws["A5"] = "Agrupación"
        ws["B5"] = granularidad.capitalize()


        ws["A6"] = "Fecha inicio"
        ws["B6"] = fecha_inicio if fecha_inicio else "Sin filtro"


        ws["A7"] = "Fecha fin"
        ws["B7"] = fecha_fin if fecha_fin else "Sin filtro"
        
        ws["A8"] = "Fecha de consulta"
        ws["B8"] = datetime.now(ZoneInfo("America/Lima")).strftime("%d/%m/%Y %H:%M")



        for fila in range(3,9):

            ws[f"A{fila}"].font = Font(
                bold=True
            )



        # ==============================
        # TABLA DE RESULTADOS
        # ==============================


        ws["A10"] = "Periodo"
        ws["B10"] = "Código servicio"
        ws["C10"] = "Servicio"
        ws["D10"] = "Total atenciones"



        for cell in ws["10:10"]:

            cell.font = Font(
                bold=True,
                color="FFFFFF"
            )

            cell.fill = PatternFill(
                "solid",
                fgColor="5E86B5"
            )



        fila = 11


        for item in data:
            ws.cell(
                fila,
                1,
                item["periodo"] if not isinstance(item.get("periodo"), str) or not re.match(r"^[=+\-@]", item["periodo"]) else f"'{item['periodo']}"
            )

            ws.cell(
                fila,
                2,
                codigo_servicio if not isinstance(codigo_servicio, str) or not re.match(r"^[=+\-@]", codigo_servicio) else f"'{codigo_servicio}"
            )

            ws.cell(
                fila,
                3,
                nombre_servicio if not isinstance(nombre_servicio, str) or not re.match(r"^[=+\-@]", nombre_servicio) else f"'{nombre_servicio}"
            )

            ws.cell(
                fila,
                4,
                item["total_atenciones"] if not isinstance(item.get("total_atenciones"), str) or not re.match(r"^[=+\-@]", str(item["total_atenciones"])) else f"'{item['total_atenciones']}"
            )

            fila += 1



        # Ajustar columnas

        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 45
        ws.column_dimensions["D"].width = 20



        # ==============================
        # GENERAR ARCHIVO
        # ==============================


        archivo = BytesIO()

        wb.save(archivo)

        archivo.seek(0)



        nombre_archivo = (
            f"Reporte_{codigo_servicio}_{codigo_establecimiento}_{granularidad}.xlsx"
        )



        return send_file(
            archivo,
            as_attachment=True,
            download_name=nombre_archivo,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )



    except Exception:

        return _error_response()


def _parametros_valorizacion():
    tipo = request.args.get("tipo", "todos").strip().lower()
    granularidad = request.args.get("granularidad", "mes").strip().lower()
    fecha_inicio = request.args.get("fecha_inicio") or None
    fecha_fin = request.args.get("fecha_fin") or None

    for valor, nombre in ((fecha_inicio, "fecha de inicio"), (fecha_fin, "fecha final")):
        if valor:
            try:
                datetime.strptime(valor, "%Y-%m-%d")
            except ValueError as exc:
                raise ValueError(f"La {nombre} no es válida.") from exc
    if fecha_inicio and fecha_fin and fecha_inicio > fecha_fin:
        raise ValueError("La fecha de inicio no puede ser posterior a la fecha final.")
    return tipo, granularidad, fecha_inicio, fecha_fin


@dashboard_bp.route("/api/valorizacion-filtrada")
@login_required
def api_valorizacion_filtrada():
    try:
        tipo, granularidad, fecha_inicio, fecha_fin = _parametros_valorizacion()
        data = kpi_service.obtener_valorizacion_filtrada(
            tipo, granularidad, fecha_inicio, fecha_fin
        )
        return jsonify(data)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    except Exception:
        return _error_response()


@dashboard_bp.route("/api/exportar-valorizacion-excel")
@login_required
def exportar_valorizacion_excel():
    try:
        tipo, granularidad, fecha_inicio, fecha_fin = _parametros_valorizacion()
        data = kpi_service.obtener_valorizacion_filtrada(
            tipo, granularidad, fecha_inicio, fecha_fin
        )
        _, tipo_label = kpi_service.TIPOS_VALORIZACION[tipo]

        wb = Workbook()
        ws = wb.active
        ws.title = "Valorización"
        ws.merge_cells("A1:E1")
        ws["A1"] = "Reporte de valorización SIS"
        ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
        ws["A1"].fill = PatternFill("solid", fgColor="17365D")

        filtros = [
            ("Tipo de valorización", tipo_label),
            ("Agrupación", granularidad.capitalize()),
            ("Fecha inicio", fecha_inicio or "Sin filtro"),
            ("Fecha fin", fecha_fin or "Sin filtro"),
            ("Fecha de consulta", datetime.now(ZoneInfo("America/Lima")).strftime("%d/%m/%Y %H:%M")),
        ]
        for fila, (etiqueta, valor) in enumerate(filtros, start=3):
            ws.cell(fila, 1, etiqueta).font = Font(bold=True)
            ws.cell(fila, 2, valor)

        encabezados = ["Periodo", "Atenciones", "Atenciones valorizadas", "Monto valorizado", "Promedio valorizado"]
        for columna, encabezado in enumerate(encabezados, start=1):
            celda = ws.cell(10, columna, encabezado)
            celda.font = Font(bold=True, color="FFFFFF")
            celda.fill = PatternFill("solid", fgColor="5E86B5")
            celda.alignment = Alignment(horizontal="center")

        for fila, item in enumerate(data, start=11):
            periodo = item.get("periodo", "")
            if isinstance(periodo, str) and re.match(r"^[=+\-@]", periodo):
                periodo = f"'{periodo}"
            valores = [periodo, item.get("total_atenciones", 0), item.get("atenciones_valorizadas", 0), item.get("monto_valorizado", 0), item.get("promedio_valorizado", 0)]
            for columna, valor in enumerate(valores, start=1):
                ws.cell(fila, columna, valor)
            ws.cell(fila, 4).number_format = '"S/ "#,##0.00'
            ws.cell(fila, 5).number_format = '"S/ "#,##0.00'

        for columna, ancho in zip("ABCDE", (18, 18, 24, 22, 22)):
            ws.column_dimensions[columna].width = ancho
        ws.freeze_panes = "A11"
        ws.auto_filter.ref = f"A10:E{max(10, 10 + len(data))}"

        archivo = BytesIO()
        wb.save(archivo)
        archivo.seek(0)
        return send_file(
            archivo,
            as_attachment=True,
            download_name=f"Valorizacion_{tipo}_{granularidad}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    except Exception:
        return _error_response()
